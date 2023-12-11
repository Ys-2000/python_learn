import requests
from bs4 import BeautifulSoup
import openpyxl
import time, random

fileName = 'gp_data.xlsx'
wb = openpyxl.load_workbook(fileName)
sheet = wb['Sheet1']


def main(page):
    url = f'https://xueqiu.com/S/{page}'

    headers = {
        "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'
    }
    try:
        response = requests.get(url, headers=headers)
        soup = BeautifulSoup(response.text, 'lxml')
        name = soup.find(class_='stock-name', ).text  # 公司名称
        try:
            intro = soup.find('div', class_='profile-detail').contents[0].text  # 简介
        except:
            intro = "无"
        # web = soup.find('div', class_='profile-detail').find('strong').text   # 公司网站
        try:
            url = soup.find('div', class_='profile-detail').find(class_="profile-link").find('a')['href']
        except:
            url = "无"
        # dza = soup.find('div', class_='profile-detail').find_all('strong')[1].text   # 公司地址:
        try:
            dz = soup.find('div', class_='profile-detail').contents[6].text   # 地址
        except:
            dz = "无"
        # tella = soup.find('div', class_='profile-detail hide').contents[-3].text      # 电话
        try:

            if soup.find('div', class_='profile-detail').contents[-2].text == "公司电话：" :
                tell = soup.find('div', class_='profile-detail').contents[-1].text  # 电话
            # elif soup.find('div', class_='profile-detail').contents[-1].text == "收起":
            #     tell = soup.find('div', class_='profile-detail').contents[-2].text  # 电话
            else:
                tell = soup.find('div', class_='profile-detail').contents[-2].text  # 电话
        except:
            tell = "无"
        # print(f"公司名称:{name}简介:{intro}官网:{url}地址:{dz}电话:{tell}")
        save(name, intro, url, dz, tell, i)
        print(f"股票代码:{page}爬取成功！！！")
        print('**' * 10)
    except Exception as e:
        print(f"发生错误：{e},{page}暂无数据'")
        print('**' * 10)


def save(name, intro, url, dz, tell, row):
    sheet.cell(row=row, column=2).value = name
    sheet.cell(row=row, column=3).value = intro
    sheet.cell(row=row, column=4).value = url
    sheet.cell(row=row, column=5).value = dz
    sheet.cell(row=row, column=6).value = tell
    wb.save(fileName)


if __name__ == '__main__':
    for i in range(2, sheet.max_row + 1):
        page = sheet[f'A{i}'].value
        main(page)
        time.sleep(random.uniform(1, 3))  # 生成 1 到 3 秒的随机延时


    # main("SH600139")